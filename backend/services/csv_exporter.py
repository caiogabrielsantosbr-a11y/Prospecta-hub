"""
CSV / Excel export utility.
"""
import io
import csv
from typing import Any
from openpyxl import Workbook


def to_csv_bytes(rows: list[dict], fields: list[str]) -> bytes:
    """Convert list of dicts to CSV bytes."""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue().encode("utf-8-sig")


def to_excel_bytes(rows: list[dict], fields: list[str], sheet_name: str = "Data") -> bytes:
    """Convert list of dicts to Excel bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header
    for col, field in enumerate(fields, 1):
        ws.cell(row=1, column=col, value=field)

    # Data
    for row_idx, row_data in enumerate(rows, 2):
        for col, field in enumerate(fields, 1):
            ws.cell(row=row_idx, column=col, value=row_data.get(field, ""))

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
